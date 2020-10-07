#!/usr/bin/env python
# encoding: gb2312

"""
@Author: yangwenhao
@Contact: 874681044@qq.com
@Software: PyCharm
@File: inputtele.py
@Time: 2020/1/29 1:24 PM
@Overview:
"""
import pandas as pd
import pathlib
import quopri
import re
from openpyxl  import load_workbook
from openpyxl.styles import Alignment

vcf_file = 'PythonProgramDesign/misc/各工地管理人员安排/00001.vcf'
contacts = {}

with open(vcf_file, 'r') as f:
    name = ''
    number = ''

    while True:
        line = f.readline()
        if line:
            # print(line)
            if line.startswith("N;CHARSET=UTF-8"):

                while not line.endswith(';;;\n'):
                    n = f.readline()
                    line += n
                try:
                    name = quopri.decodestring(line.split(':')[-1].rstrip('\n')).decode(errors='ignore')
                    name = re.sub('[;]', '', name)
                    print(name)
                except:
                    print(line)

                continue

            elif line.startswith("TEL"):
                number = line.split(':')[-1]
                continue

            elif line.startswith("END:VCARD"):
                contacts[name]=number
                name=''
                number=''
        else:
            break

contactors = list(contacts.keys())
for k in contacts:
    v=contacts[k]
    contacts[k]=re.sub('[ |\n]', '', v)

file_dir = 'PythonProgramDesign/misc/各工地管理人员安排'
exc = pathlib.Path(file_dir)
exc_list = list(exc.glob('*.xlsx'))

for e in exc_list:

    b2 = load_workbook(str(e))
    print(str(exc_list[0]))
    sh1 = b2[b2.sheetnames[0]]

    align = Alignment(horizontal='center', vertical='center', wrapText=True)

    for i in range(1, 46):
        if sh1['B%d' % i].value != None:
            name = sh1['B%d' % i].value
            name = re.sub('[\b]', '', name)

            for c in contactors:
                if name in c:
                    if sh1['C%d' % i].value != None:
                        ori_va = sh1['C%d' % i].value + ','
                    else:
                        ori_va = ''
                    sh1['C%d' % i].value = ori_va + contacts[c]
                    sh1['C%d' % i].alignment = align

    b2.save(str(e))


